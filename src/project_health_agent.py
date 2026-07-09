#!/usr/bin/env python3
"""Project health reporting agent.

Reads one or more Excel project plans, extracts health signals, assigns RAG
status, and writes weekly plain-English reports plus a portfolio JSON summary.
"""

from __future__ import annotations

import argparse
import json
import math
import re
from collections import Counter
from dataclasses import dataclass, asdict
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Iterable

import openpyxl


MISSING = {"", "none", "null", "nan", "#unparseable"}
DONE_STATUSES = {"completed", "complete", "not applicable", "n/a", "na"}
ACTIVE_STATUSES = {"not started", "in progress", "on hold", "blocked"}
NEGATIVE_TERMS = {
    "delay",
    "delayed",
    "impacted",
    "pending",
    "blocked",
    "risk",
    "issue",
    "miss",
    "missed",
    "need",
    "remain",
    "rework",
    "dependency",
    "waiting",
}
POSITIVE_TERMS = {"covered", "complete", "completed", "aligned", "signed", "approved"}


def normalize(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value or "").strip().lower())


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    text = str(value).strip()
    return "" if text.lower() in MISSING else text


def parse_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)) and not math.isnan(value):
        try:
            return openpyxl.utils.datetime.from_excel(value).date()
        except Exception:
            return None
    if isinstance(value, str):
        text = value.strip()
        for fmt in ("%Y-%m-%d", "%m/%d/%y %I:%M %p", "%m/%d/%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                pass
    return None


def truthy(value: Any) -> bool:
    text = clean_text(value).lower()
    return value is True or text in {"true", "yes", "y", "1", "high", "red", "at risk"}


def color_level(value: Any) -> str | None:
    text = clean_text(value).lower()
    if text in {"red", "r"}:
        return "Red"
    if text in {"yellow", "amber", "a"}:
        return "Amber"
    if text in {"green", "g"}:
        return "Green"
    return None


def level_value(level: str) -> int:
    return {"Green": 0, "Amber": 1, "Red": 2}.get(level, 0)


def find_col(headers: list[Any], aliases: Iterable[str]) -> int | None:
    normalized = [normalize(h) for h in headers]
    alias_set = {normalize(a) for a in aliases}
    for idx, header in enumerate(normalized):
        if header in alias_set:
            return idx
    for idx, header in enumerate(normalized):
        if any(alias in header for alias in alias_set if alias):
            return idx
    return None


def ratio(part: int, whole: int) -> float:
    return part / whole if whole else 0.0


@dataclass
class ProjectHealth:
    source_file: str
    project_name: str
    project_manager: str
    project_stage: str
    as_of: str
    rag: str
    confidence: str
    reason: str
    recommendations: list[str]
    signals: dict[str, Any]
    metrics: dict[str, Any]
    evidence: dict[str, list[str]]


class ProjectHealthAgent:
    def __init__(self, as_of: date, config: dict[str, Any] | None = None):
        self.as_of = as_of
        self.weights = (config or {}).get(
            "weights",
            {
                "schedule": 0.35,
                "milestones": 0.20,
                "blockers": 0.20,
                "progress": 0.15,
                "stakeholder_sentiment": 0.05,
                "data_quality": 0.05,
            },
        )
        self.thresholds = (config or {}).get(
            "thresholds",
            {
                "red_score": 1.20,
                "amber_score": 0.55,
            },
        )

    def evaluate(self, workbook_path: Path) -> ProjectHealth:
        wb = openpyxl.load_workbook(workbook_path, data_only=True, read_only=True)
        plan_ws = self._find_plan_sheet(wb)
        headers = [cell for cell in next(plan_ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        cols = self._map_columns(headers)
        raw_rows = list(plan_ws.iter_rows(min_row=2, values_only=True))
        rows = [self._row_dict(row, cols) for row in raw_rows]
        rows = [row for row in rows if any(row.get(k) for k in ("task", "status", "schedule_health"))]
        summary = self._read_summary(wb)
        comments = self._read_comments(wb)

        metrics = self._metrics(rows, summary, comments)
        signals = self._signals(metrics, summary)
        rag = self._overall_rag(signals)
        reason = self._reason(rag, metrics, signals)
        recommendations = self._recommendations(metrics, signals)

        project_name = (
            clean_text(summary.get("Project Name"))
            or self._first_value(rows, "project_name")
            or self._first_task_name(rows)
            or workbook_path.stem
        )
        project_manager = clean_text(summary.get("Project Manager")) or self._first_value(rows, "manager") or "Unknown"
        project_stage = clean_text(summary.get("Project Stage")) or metrics["current_stage"] or "Not specified"

        return ProjectHealth(
            source_file=str(workbook_path),
            project_name=project_name,
            project_manager=project_manager,
            project_stage=project_stage,
            as_of=self.as_of.isoformat(),
            rag=rag,
            confidence=self._confidence(metrics),
            reason=reason,
            recommendations=recommendations,
            signals=signals,
            metrics=metrics,
            evidence={
                "overdue_examples": [x["task"] for x in metrics["overdue_examples"]],
                "upcoming_examples": [x["task"] for x in metrics["upcoming_examples"]],
                "blocker_comments": metrics["blocker_comments"][:5],
                "data_quality_notes": metrics["data_quality_notes"],
            },
        )

    def _find_plan_sheet(self, wb: Any) -> Any:
        candidates = [ws for ws in wb.worksheets if ws.title.lower() not in {"summary", "comments"}]
        return max(candidates or wb.worksheets, key=lambda ws: ws.max_row * ws.max_column)

    def _map_columns(self, headers: list[Any]) -> dict[str, int | None]:
        return {
            "project_name": find_col(headers, ["Project Name"]),
            "manager": find_col(headers, ["Project Manager"]),
            "phase": find_col(headers, ["Phase/Milestone", "Milestone", "Phase"]),
            "level": find_col(headers, ["Level", "Ancestors"]),
            "risk": find_col(headers, ["At Risk?", "At Risk"]),
            "schedule_health": find_col(headers, ["Schedule Health"]),
            "task": find_col(headers, ["Task Name"]),
            "status": find_col(headers, ["Status"]),
            "start": find_col(headers, ["Start Date"]),
            "end": find_col(headers, ["End Date"]),
            "percent_complete": find_col(headers, ["% Complete", "Percent Complete"]),
            "on_hold": find_col(headers, ["On Hold?"]),
            "critical": find_col(headers, ["Critical ?","Critical?","Critical"]),
            "baseline_start": find_col(headers, ["Baseline Start", "Baseline Start Date"]),
            "baseline_finish": find_col(headers, ["Baseline Finish", "Baseline End Date"]),
            "variance": find_col(headers, ["Variance"]),
            "status_comment": find_col(headers, ["Status Comment", "Comments"]),
            "owner": find_col(headers, ["Owner", "Assigned To"]),
        }

    def _row_dict(self, row: tuple[Any, ...], cols: dict[str, int | None]) -> dict[str, Any]:
        def get(name: str) -> Any:
            idx = cols.get(name)
            return row[idx] if idx is not None and idx < len(row) else None

        return {
            "project_name": clean_text(get("project_name")),
            "manager": clean_text(get("manager")),
            "phase": clean_text(get("phase")),
            "level": get("level"),
            "risk": truthy(get("risk")),
            "schedule_health": color_level(get("schedule_health")),
            "task": clean_text(get("task")),
            "status": clean_text(get("status")),
            "start": parse_date(get("start")),
            "end": parse_date(get("end")),
            "percent_complete": get("percent_complete") if isinstance(get("percent_complete"), (int, float)) else None,
            "on_hold": truthy(get("on_hold")),
            "critical": truthy(get("critical")),
            "baseline_finish": parse_date(get("baseline_finish")),
            "variance": get("variance") if isinstance(get("variance"), (int, float)) else None,
            "status_comment": clean_text(get("status_comment")),
            "owner": clean_text(get("owner")),
        }

    def _read_summary(self, wb: Any) -> dict[str, Any]:
        if "Summary" not in wb.sheetnames:
            return {}
        summary = {}
        for key, value, *_ in wb["Summary"].iter_rows(values_only=True):
            if clean_text(key):
                summary[clean_text(key)] = value
        return summary

    def _read_comments(self, wb: Any) -> list[str]:
        if "Comments" not in wb.sheetnames:
            return []
        comments = []
        for row in wb["Comments"].iter_rows(values_only=True):
            values = [clean_text(v) for v in row if clean_text(v)]
            if values:
                comments.append(" | ".join(values))
        return comments

    def _metrics(self, rows: list[dict[str, Any]], summary: dict[str, Any], comments: list[str]) -> dict[str, Any]:
        active_rows = [r for r in rows if r["status"].lower() not in DONE_STATUSES]
        completed = sum(1 for r in rows if r["status"].lower() in {"completed", "complete"})
        overdue = [r for r in active_rows if r["end"] and r["end"] < self.as_of]
        upcoming = [r for r in active_rows if r["end"] and self.as_of <= r["end"] <= self.as_of + timedelta(days=14)]
        schedule_counts = Counter(r["schedule_health"] for r in rows if r["schedule_health"])
        status_counts = Counter(r["status"] for r in rows if r["status"])
        risk_rows = [r for r in rows if r["risk"]]
        on_hold = [r for r in active_rows if r["on_hold"] or r["status"].lower() == "on hold"]
        red_active = [r for r in active_rows if r["schedule_health"] == "Red"]
        amber_active = [r for r in active_rows if r["schedule_health"] == "Amber"]
        milestones = [r for r in rows if r["phase"] or self._is_top_level(r)]
        active_milestones = [r for r in milestones if r["status"].lower() not in DONE_STATUSES]
        overdue_milestones = [r for r in active_milestones if r["end"] and r["end"] < self.as_of]
        red_milestones = [r for r in active_milestones if r["schedule_health"] == "Red"]

        negative_comments = []
        positive_comment_count = 0
        for text in comments + [r["status_comment"] for r in rows if r["status_comment"]]:
            lowered = text.lower()
            if any(term in lowered for term in NEGATIVE_TERMS):
                negative_comments.append(text)
            if any(term in lowered for term in POSITIVE_TERMS):
                positive_comment_count += 1

        start = parse_date(summary.get("Project Start Date")) or self._min_date(r["start"] for r in rows)
        end = parse_date(summary.get("Project End Date")) or self._max_date(r["end"] for r in rows)
        pct_summary = summary.get("% Complete")
        pct_complete = pct_summary if isinstance(pct_summary, (int, float)) else ratio(completed, len(rows))
        elapsed_ratio = ratio((self.as_of - start).days, (end - start).days) if start and end and end > start else None
        progress_gap = (elapsed_ratio - pct_complete) if elapsed_ratio is not None else None

        data_quality_notes = []
        if any("#UNPARSEABLE" in str(cell) for r in rows[:20] for cell in r.values()):
            data_quality_notes.append("Workbook contains unparseable formula/export values; ignored for core health scoring.")
        if not any("Budget" in clean_text(k) for k in summary):
            data_quality_notes.append("Budget burn data is not present; budget signal is treated as neutral/missing.")
        if not comments:
            data_quality_notes.append("No stakeholder comments were available; sentiment confidence is lower.")

        return {
            "total_tasks": len(rows),
            "active_tasks": len(active_rows),
            "status_counts": dict(status_counts),
            "schedule_counts": dict(schedule_counts),
            "red_active_tasks": len(red_active),
            "amber_active_tasks": len(amber_active),
            "overdue_active_tasks": len(overdue),
            "upcoming_14_day_active_tasks": len(upcoming),
            "risk_flagged_tasks": len(risk_rows),
            "on_hold_tasks": len(on_hold),
            "milestone_count": len(milestones),
            "active_milestones": len(active_milestones),
            "overdue_milestones": len(overdue_milestones),
            "red_milestones": len(red_milestones),
            "percent_complete": round(pct_complete, 4) if isinstance(pct_complete, (int, float)) else None,
            "time_elapsed_ratio": round(elapsed_ratio, 4) if elapsed_ratio is not None else None,
            "progress_gap": round(progress_gap, 4) if progress_gap is not None else None,
            "summary_at_risk": clean_text(summary.get("At Risk")),
            "summary_schedule_health": color_level(summary.get("Schedule Health")) or clean_text(summary.get("Schedule Health")),
            "current_stage": clean_text(summary.get("Project Stage")),
            "negative_comment_count": len(negative_comments),
            "positive_comment_count": positive_comment_count,
            "blocker_comments": negative_comments,
            "overdue_examples": self._examples(overdue),
            "upcoming_examples": self._examples(upcoming),
            "data_quality_notes": data_quality_notes,
        }

    def _signals(self, metrics: dict[str, Any], summary: dict[str, Any]) -> dict[str, Any]:
        active = metrics["active_tasks"]
        red_ratio = ratio(metrics["red_active_tasks"], active)
        overdue_ratio = ratio(metrics["overdue_active_tasks"], active)
        summary_schedule = metrics["summary_schedule_health"]

        schedule = "Green"
        if summary_schedule == "Red" or red_ratio >= 0.08 or overdue_ratio >= 0.18:
            schedule = "Red"
        elif summary_schedule == "Amber" or red_ratio > 0 or overdue_ratio > 0:
            schedule = "Amber"

        progress = "Green"
        gap = metrics["progress_gap"]
        if gap is not None:
            if gap >= 0.25:
                progress = "Red"
            elif gap >= 0.10:
                progress = "Amber"

        milestone = "Green"
        if metrics["red_milestones"] > 0 or metrics["overdue_milestones"] > 0:
            milestone = "Red"
        elif metrics["upcoming_14_day_active_tasks"] > 0:
            milestone = "Amber"

        blockers = "Green"
        if metrics["summary_at_risk"].lower() == "high" or metrics["on_hold_tasks"] > 0:
            blockers = "Red"
        elif metrics["risk_flagged_tasks"] > 0 or metrics["negative_comment_count"] > 0:
            blockers = "Amber"

        sentiment = "Green"
        if metrics["negative_comment_count"] >= 3:
            sentiment = "Red"
        elif metrics["negative_comment_count"] > 0:
            sentiment = "Amber"
        elif not metrics["blocker_comments"]:
            sentiment = "Neutral"

        return {
            "schedule": schedule,
            "progress": progress,
            "milestones": milestone,
            "blockers": blockers,
            "stakeholder_sentiment": sentiment,
            "budget": "Missing",
            "data_quality": "Amber" if metrics["data_quality_notes"] else "Green",
        }

    def _overall_rag(self, signals: dict[str, Any]) -> str:
        score = sum(level_value(signals.get(k, "Green")) * w for k, w in self.weights.items())
        if signals["schedule"] == "Red" and signals["blockers"] in {"Amber", "Red"}:
            return "Red"
        if score >= self.thresholds.get("red_score", 1.20):
            return "Red"
        if score >= self.thresholds.get("amber_score", 0.55):
            return "Amber"
        return "Green"

    def _reason(self, rag: str, metrics: dict[str, Any], signals: dict[str, Any]) -> str:
        parts = [
            f"Overall status is {rag} because schedule is {signals['schedule']} and blockers are {signals['blockers']}.",
            f"{metrics['overdue_active_tasks']} active tasks are past their planned end date, with {metrics['red_active_tasks']} active tasks already marked Red by the plan.",
        ]
        if metrics["progress_gap"] is not None:
            gap_pp = round(metrics["progress_gap"] * 100, 1)
            if gap_pp > 0:
                parts.append(f"Progress is about {gap_pp} percentage points behind elapsed time.")
            else:
                parts.append(f"Progress is ahead of elapsed time by about {abs(gap_pp)} percentage points.")
        if metrics["summary_at_risk"]:
            parts.append(f"The workbook summary marks project risk as {metrics['summary_at_risk']}.")
        if metrics["negative_comment_count"]:
            parts.append(f"Stakeholder/commentary evidence includes {metrics['negative_comment_count']} blocker-like notes.")
        return " ".join(parts)

    def _recommendations(self, metrics: dict[str, Any], signals: dict[str, Any]) -> list[str]:
        recs = []
        if signals["schedule"] == "Red":
            recs.append("Run a two-week recovery plan for overdue Red and overdue active tasks, with named owners and daily closure tracking.")
        if signals["milestones"] == "Red":
            recs.append("Reconfirm milestone dates and publish a revised baseline for any milestone already past due.")
        if signals["blockers"] in {"Amber", "Red"}:
            recs.append("Escalate open dependencies and at-risk items in the weekly client steering forum.")
        if signals["budget"] == "Missing":
            recs.append("Add budget planned, actual, and forecast fields so financial burn can be scored instead of treated as missing.")
        if metrics["upcoming_14_day_active_tasks"]:
            recs.append(f"Pre-review the {metrics['upcoming_14_day_active_tasks']} active tasks due in the next 14 days to prevent additional slippage.")
        return recs[:5]

    def _confidence(self, metrics: dict[str, Any]) -> str:
        if metrics["total_tasks"] and metrics["summary_schedule_health"]:
            return "High"
        if metrics["total_tasks"]:
            return "Medium"
        return "Low"

    def _examples(self, rows: list[dict[str, Any]], limit: int = 8) -> list[dict[str, str]]:
        rows = sorted(rows, key=lambda r: (r["end"] or date.max, r["schedule_health"] != "Red"))
        return [
            {
                "task": r["task"],
                "status": r["status"],
                "schedule_health": r["schedule_health"] or "",
                "end": r["end"].isoformat() if r["end"] else "",
                "phase": r["phase"],
                "owner": r["owner"],
            }
            for r in rows[:limit]
        ]

    def _is_top_level(self, row: dict[str, Any]) -> bool:
        level = row.get("level")
        if isinstance(level, (int, float)):
            return level in {0, 1}
        return False

    def _first_value(self, rows: list[dict[str, Any]], key: str) -> str:
        return next((clean_text(r.get(key)) for r in rows if clean_text(r.get(key))), "")

    def _first_task_name(self, rows: list[dict[str, Any]]) -> str:
        return next((r["task"] for r in rows if r["task"]), "")

    def _min_date(self, values: Iterable[date | None]) -> date | None:
        dates = [v for v in values if v]
        return min(dates) if dates else None

    def _max_date(self, values: Iterable[date | None]) -> date | None:
        dates = [v for v in values if v]
        return max(dates) if dates else None


def write_markdown_report(result: ProjectHealth, out_dir: Path) -> Path:
    safe_name = re.sub(r"[^A-Za-z0-9]+", "_", Path(result.source_file).stem).strip("_")
    path = out_dir / f"{safe_name}_weekly_report.md"
    m = result.metrics
    lines = [
        f"# Weekly Project Health Report: {result.project_name}",
        "",
        f"**As of:** {result.as_of}",
        f"**Project Manager:** {result.project_manager}",
        f"**Stage:** {result.project_stage}",
        f"**RAG:** {result.rag}",
        f"**Confidence:** {result.confidence}",
        "",
        "## Executive Reasoning",
        result.reason,
        "",
        "## Signal Breakdown",
    ]
    for key, value in result.signals.items():
        lines.append(f"- {key.replace('_', ' ').title()}: {value}")
    lines.extend([
        "",
        "## Key Metrics",
        f"- Total tasks: {m['total_tasks']}",
        f"- Active tasks: {m['active_tasks']}",
        f"- Active overdue tasks: {m['overdue_active_tasks']}",
        f"- Active tasks due in next 14 days: {m['upcoming_14_day_active_tasks']}",
        f"- Active Red tasks: {m['red_active_tasks']}",
        f"- Active Amber tasks: {m['amber_active_tasks']}",
        f"- At-risk flags: {m['risk_flagged_tasks']}",
        f"- On-hold tasks: {m['on_hold_tasks']}",
        f"- Percent complete: {round((m['percent_complete'] or 0) * 100, 1)}%",
        "",
        "## Recommended Actions",
    ])
    lines.extend([f"- {rec}" for rec in result.recommendations])
    lines.extend(["", "## Evidence", "### Overdue Examples"])
    for item in m["overdue_examples"]:
        lines.append(f"- {item['end']} | {item['schedule_health'] or 'No health'} | {item['status']} | {item['task']}")
    lines.append("")
    lines.append("### Upcoming Examples")
    for item in m["upcoming_examples"]:
        lines.append(f"- {item['end']} | {item['schedule_health'] or 'No health'} | {item['status']} | {item['task']}")
    if result.evidence["blocker_comments"]:
        lines.extend(["", "### Blocker-Like Comments"])
        lines.extend([f"- {comment}" for comment in result.evidence["blocker_comments"]])
    if result.evidence["data_quality_notes"]:
        lines.extend(["", "### Data Quality Notes"])
        lines.extend([f"- {note}" for note in result.evidence["data_quality_notes"]])
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return path


def main() -> int:
    parser = argparse.ArgumentParser(description="Generate project health reports from Excel plans.")
    parser.add_argument("plans", nargs="+", type=Path, help="Excel project plan files")
    parser.add_argument("--as-of", default=date.today().isoformat(), help="Assessment date, YYYY-MM-DD")
    parser.add_argument("--output-dir", type=Path, default=Path("sample_outputs"))
    args = parser.parse_args()

    as_of = datetime.strptime(args.as_of, "%Y-%m-%d").date()
    args.output_dir.mkdir(parents=True, exist_ok=True)
    agent = ProjectHealthAgent(as_of)
    results = []
    report_paths = []
    for plan in args.plans:
        result = agent.evaluate(plan)
        results.append(asdict(result))
        report_paths.append(str(write_markdown_report(result, args.output_dir)))

    summary_path = args.output_dir / "project_health_summary.json"
    summary_path.write_text(json.dumps({"projects": results}, indent=2, default=str), encoding="utf-8")
    print(json.dumps({"summary": str(summary_path), "reports": report_paths}, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
